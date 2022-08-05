#Load app and configuration
# create config variables (to be cleaned in the future)

from flasky import db
from config import config as config_set
config=config_set['tinymrp'].__dict__


from .models import Part, Bom , solidbom, thumbnail

import pandas as pd


from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import openpyxl
import os




#This function provides a flatbom where only parts and welded assemblies 
## are listed excluding fasteners by default



def bom_to_excel(flatbom,outputfolder,title=""):
    
        sheet="Manufactured components"
 
        if title=="": title="List"
        
        excel_file=outputfolder + title +".xlsx"
        
        property_list=['partnumber','revision','qty','description','material','thickness','finish']

        
        # for pepe in flatbom:
            #print(pepe)
        
        paco=[x.as_dict() for x in flatbom]
        bom_in=pd.DataFrame([x.as_dict() for x in flatbom])
        bom_to_sheet=bom_in[property_list] 
        
        
        #Generate basic lists:
        with pd.ExcelWriter(excel_file) as writer:
            bom_to_sheet.to_excel(writer, sheet_name=sheet)
            
            


 
        
        workbook=openpyxl.load_workbook(filename=excel_file)
        # # #print(dir(self.workbook))
        
        # Add images
        worksheet=workbook[sheet]
        worksheet.insert_cols(idx=2)
        

        
        #Get the a
        str_flatbom=bom_to_sheet.applymap(str)
        col_width=[]

        for col in str_flatbom:
           col_width.append(str_flatbom[col].map(len).max())
        col_width.insert(0,5)
        col_width.insert(1,10)

        
        
        #print(col_width)
       
        for i, column_width in enumerate(col_width):
            if column_width>5:
                worksheet.column_dimensions[get_column_letter(i+1)].width =int( 1.1*column_width)


        
                

        i=-1           
        for index, row in bom_in.iterrows():
            i=i+1
            rd =worksheet.row_dimensions[i+2] # get dimension for row 3
            rd.height = 40 # value in points, there is no "auto"
            png_folder=config['DELIVERABLES_FOLDER']+"png/"
            thumb=thumbnail(png_folder+row.at['file']+"_REV_"+row['revision']+'.png',
                            size=(100, 100))
    
               
            #Add image
            cell='B'+str(i+2)
            try:
                image=openpyxlIm(thumb)
                image.height=60
                image.width=60
                worksheet.add_image(image, cell)
                
            except:
                # worksheet.add_image(thumb, cell)
                #print("Could not add image to excel ", row['partnumber'])
                pass
                
            
        
        #create the border style to put all around the cells
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))   

        #Border annd Center alignment for all cells    
        for row in worksheet.iter_rows():
            for cell in row:
                # #print(dir(cell.style))
                # cell.style.alignment.wrap_text=True
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        #Put to landscape and adjust the width of the page to the width of the content
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(worksheet, paper_size = int(sum(col_width)), orientation='landscape')
        workbook.save(filename=excel_file)
        
        #Export to pdf
        pdf_path=outputfolder + title +".pdf"
        #excel_to_pdf (excel_file, pdf_path,erasepdf=True)
        
        return excel_file
