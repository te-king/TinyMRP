#Load app and configuration
# create config variables (to be cleaned in the future)

from flasky import db
from flask_login import login_required, current_user

from config import config as config_set

from app.models import User

config=config_set['tinymrp'].__dict__

folderout=config['FOLDEROUT']
fileserver_path=config['FILESERVER_PATH']
datasheet_folder=config['DATASHEET_FOLDER']
deliverables_folder=config['DELIVERABLES_FOLDER']
variables_conf=config['VARIABLES_CONF']
webfileserver=config['WEBFILESERVER']
maincols=config['MAINCOLS']
refcols=config['REFCOLS']
deliverables=config['DELIVERABLES']
webserver=config['WEBSERVER']
process_conf=config['PROCESS_CONF']
lowercase_properties=config['LOWERCASE_PROPERTIES']
property_conf=config['PROPERTY_CONF']
hardware_folder=config['HARDWARE_FOLDER']



from flask import (
    Blueprint, flash, g, redirect, session, render_template, request, url_for
)


#To genearte qr codes
import qrcode



#Other libraries
from datetime import datetime, date #for timestamps

import chardet #for getting the encoding of files
import sys, os
import re
from shutil import copyfile
import glob
import pickle #To save bom object session
from pathlib import Path, PureWindowsPath, PurePosixPath
import pandas as pd



import numpy as np
import re
import math



#PDF libraries
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
import pdfkit #TO EXPORT WEBPAGES TO PDF


# import openpyxl #to manipulate excelfiles
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as openpyxlIm
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from PIL import Image #to process thumbnails


#SQL libraries
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from sqlalchemy.types import Integer, Float, Boolean,String,Text,NVARCHAR, Date
from flask import render_template, jsonify, request, redirect, url_for, jsonify
from sqlalchemy import create_engine, ForeignKey,select, or_, and_
from sqlalchemy.orm import relationship, backref




#To check if conection to file.
import urllib



#Tinylib internal imports
from .imageprocess import cropandbackground



#Mongorelated
#from mongoengine import *
from mongoengine import EmbeddedDocument,EmbeddedDocumentField, \
                        StringField,ListField,IntField, DynamicDocument,\
                        ReferenceField,DynamicField
import pymongo
import json
from pymongo import MongoClient
from werkzeug.utils import secure_filename

client = pymongo.MongoClient("localhost", 27017)
mongodb=client.TinyMRP
partcol=mongodb["part"]




#Core inventory class 
class mongoPart(DynamicDocument):
    meta = {'collection': 'part'}
    # partnumber=StringField(required=True)
    partnumber=StringField()
    description=StringField()
    revision=StringField()
    finish=StringField()
    children=ListField(ReferenceField("self"))
    childrenqty=ListField(IntField())


    #Uploader
    approved=StringField()
    uploader=DynamicField()
    #Procurement/customer
    category=DynamicField()
    supplier=DynamicField()
    customer=DynamicField()
    
    #Related files locations
    modelpath=StringField()
    pngpath=StringField()
    pdfpath=StringField()
    dxfpath=StringField()
    edrpath=StringField()
    edr_dpath=StringField()
    steppath=StringField()
    threemfpath=StringField()
    datasheetpath=StringField()
    qrpath=StringField()
    #Real pictures of the object
    picpath=ListField(StringField())
    #pics for drawings allows several pages (several images)
    png_dpath=StringField()
    #For another attachements
    otherspath=ListField(StringField())
    
    #Processes
    process=DynamicField()
    process_colors=DynamicField()
    process_icons=DynamicField()

    tag=StringField()



    def hasConsumingProcess(self):
        consume=False

        processlist=self.process
        processlist=[x for x in processlist if x in process_conf.keys()]

        for process in processlist:
            if int(process_conf[process]['priority'])<20:
                consume=True

        return consume



    def isMainProcess(self,process):
        mainprocess_bool=False

        if self.hasProcess(process):
            processlist=self.process
            processlist=[x for x in processlist if x in process_conf.keys()]
            if len(processlist)>0:
                for x in processlist:
                    if int(process_conf[process]['priority'])<= int(process_conf[x]['priority']):
                        mainprocess_bool=True
                        
                    else:
                        mainprocess_bool=False
            

        return mainprocess_bool


    def MainProcess(self):
        processlist=self.process
        print(self.partnumber)
        print(processlist)
        if len(processlist)>0: processlist=[x for x in processlist if x in process_conf.keys()]

        
        if len(processlist)>0:
            mainprocess=processlist[0]
            for x in processlist:
                if int(process_conf[x]['priority'])<= int(process_conf[mainprocess]['priority']):
                    mainprocess=x

        else:
            mainprocess='other'
        self.mainprocess=mainprocess
                
        return mainprocess

    
    #Get tree for tree representation:
    def treeDict(self, count=0,qty=1):
         
        
        refkids=[]
        i=0
        for child in self.children:
            kid={}
            count+=1
            kid=mongoPart.objects(pk=self.children[i]['id']).first().to_dict()
            kid['qty']=self.childrenqty[i]
            kid['value']=kid['qty']
            kid['children']=self.children[i].treeDict(count=count)['children']
            kid['name']=self.children[i].treeDict(count=count)['partnumber']
            refkids.append(kid)
            i+=1
        
        refdict=self.to_dict()
        refdict['children']=refkids
        refdict['name']=self.partnumber
        try:
            refdict['value']=self.qty
        except:
            refdict['value']=qty
        

        #test['children']['0']['part']['children']

        # refkids=[]
        # for kid in refdict['children']:
        #     kid['part']['children']=mongoPart.objects(id=kid['part']['id']).first().treeDict()
        #     print(kid)
        #     refkids.append(kid)
        #     # refkids.append({'part': mongoPart.objects(id=t).first().to_dict())  
        # refdict['children']=refkids

        return refdict

        


    #For the print outs on terminal of the object
    def __repr__(self):
        return f'P/N_{self.partnumber}_REV_{self.revision}_DES_{self.description}'
    def __str__ (self):
        return f'P/N_{self.partnumber}_REV_{self.revision}_DES_{self.description}'
    

    #To get the children with quantities
    def getchildren (self):
        outlist=[]
        i=0
        for child in self.children:
            kid={}
            # kid['part']=partcol.find_one({"_id":child.id})
            kid['part']=mongoPart.objects(pk=child.id)[0]
            kid['qty']=self.childrenqty[i]
            outlist.append(kid)
            i+=1
        return outlist

    def children_with_qty (self):
        outlist=[]
        i=0
        for child in self.children:
            kid=child
            # kid['part']=partcol.find_one({"_id":child.id})
            kid['part']=mongoPart.objects(pk=child.id)[0]
            kid['qty']=self.childrenqty[i]
            outlist.append(kid)
            i+=1
        return outlist
    

    #To get the parents with quantities in those parents
    def getparents (self):
        outlist=[]
        i=0
        parents=mongoPart.objects(children=self.pk)
        
        i=0
        for parent in parents:
            father={}
            father['part']=parent
            j=0
            for kid in parent.children:
                if kid.pk==self.pk:
                    father['qty']=parent.childrenqty[j]
                j+=1
            outlist.append(father)
            i+=1
               
        return outlist

    def parents_with_qty(self):
        outlist=[]
      
        parents=mongoPart.objects(children=self.pk)
    
        for parent in parents:
            j=0
            for kid in parent.children:
                if kid.pk==self.pk:
                    parent['qty']=parent.childrenqty[j]
                j+=1
            outlist.append(parent)
            
               
        return outlist

        
    def get_process_icons (self,persist=False):
        self['process_icons']=[]
        self['process_colors']=[]

        if not "process2" in self.to_dict().keys(): self['process2']=""
        if not "process3" in self.to_dict().keys(): self['process3']=""

        if type(self['process'])==str:
            self['process']=[self['process']]
            if self['process2'] and self['process2']!="" and self['process2']!=" ":
                 self['process'].append(self['process2'])
            if self['process3'] and self['process3']!="" and self['process3']!=" ":
                 self['process'].append(self['process3'])
            persist=True 

        if type(self['process'])==list:
            
            self['process']=[x for x in self['process'] if x and x!="" and x!=" "]
            
            #Put all to lowercase
            self['process']=list(map(lambda x: x.lower(), self['process']))

            #Remove duplicates
            self['process']=list(dict.fromkeys(self['process']))

            for process in self['process']:
                if process in process_conf.keys() :
                    self['process_icons'].append('images/'+(process_conf[process]['icon']))
                    self['process_colors'].append(process_conf[process]['color'])
                else:
                    self['process_icons'].append('images/'+process_conf['others']['icon'])
                    self['process_colors'].append(process_conf['others']['color'])
        
        if persist: self.save() 

       
    #To return a dictinary with the parts attributes
    def to_dict(self):
        print(self.partnumber)
        
        # dirtydict=self.to_mongo()#.to_dict()
        #

        dirtydict={}
        dirtydict['partnumber']=self.partnumber
        dirtydict['revision']=self.revision
        dirtydict['description']=self.description
        dirtydict['process']=self.process
        dirtydict['finish']=self.finish
        dirtydict['children']=self.children
        dirtydict['pngpath']=self.pngpath

        
        dirtydict=self.to_mongo().to_dict()
        dirtydict['_id']=str(dirtydict['_id'])
        
        print(dirtydict)
        
        cleanchildren=[]
        for child in dirtydict['children']:
            cleanchildren.append(str(child))
        
        dirtydict['children']=cleanchildren

        

        return dirtydict
    
    #Generates a time stamp tag with partnumber and rev
    def get_tag(self):
        self.tag=self.partnumber+"_REV_"+self.revision+"-"+date.today().strftime('%d_%m_%Y')
        return self.tag

    #Checks if a process is present
    def hasProcess(self,process):
        if process in self.process:
            return True
        else:
            return False


    #Checks if it has a consuming process
    #That is to say a special process that should hide
    #the children as they are subcomponents that cannot
    #be supplied without creating the parent
    def hasConsumingProcess(self):
        processlist=[x for x in self.process if x in process_conf.keys()]
        consume=False
        for process in processlist:
            if int(process_conf[process]['priority'])<20:
                consume=True
        return consume

    #Checks if 
    def isMainProcess(self,process):
        mainprocess_bool=False

        if self.hasProcess(process):
            processlist=[self.process,self.process2,self.process3]
            processlist=[x for x in processlist if x in process_conf.keys()]
            if len(processlist)>0:
                for x in processlist:
                    if int(process_conf[process]['priority'])<= int(process_conf[x]['priority']):
                        mainprocess_bool=True
                        
                    else:
                        mainprocess_bool=False
            

        return mainprocess_bool

    

    #Check for all available files in the Fileserver folders for a part
    def updateFileset(self,web=False,persist=False):
        self.get_tag()
        parttag=self.partnumber+"_REV_"+self.revision
        save=False
        for filetype in config['DELIVERABLES']:
            filelist=[]
            for extension in config['DELIVERABLES'][filetype]['extension']:
                
                filetag=config['DELIVERABLES'][filetype]['path']+parttag+str(config['DELIVERABLES'][filetype]['filemod'])+"."+extension
                
                if file_exists(filetag):
                    if config['DELIVERABLES'][filetype]['list']!="yes":
                        try:
                            if self[filetype+'path']!=filetag:
                                self[filetype+'path']=filetag
                                # print(filetype,extension)
                                # print("string- " ,filetag)
                                save=True
                        except:
                            print("couldnt save - ", "string- " ,filetag)
                    else:
                        if self[filetype+'path']==[] or self[filetype+'path']==None:
                            self[filetype+'path']=[]

                        if not ( filetag in self[filetype+'path'] ):
                            self[filetype+'path'].append(filetag)
                            # print(filetype,extension)
                            # print("list- ", filetag)
                            save=True
                else:
                    pass
                    # if filetype+'path'=='pngpath':
                    #     self[filetype+'path']=url_for('static', filename='images/logo.png')
                    #     save=True


                if web:
                    try:
                        self[filetype+'path']=self[filetype+'path'].replace(fileserver_path,webfileserver)

                        
                        
                        print( (self[filetype+'path']))
                        print( secure_filename(self[filetype+'path']))
                    except:
                        pass
        if persist:
            if save: self.save()


    def getweblinks(self,checkfiles=False):
        if self.partnumber ==None:
            pass
        else:
            parttag=self.partnumber+"_REV_"+self.revision

            if checkfiles:
                self.updateFileset(web=True)
            else:        
                for filetype in config['DELIVERABLES']:
                    filelist=[]
                    for extension in config['DELIVERABLES'][filetype]['extension']:
                        filetag=config['DELIVERABLES'][filetype]['path']+parttag+str(config['DELIVERABLES'][filetype]['filemod'])+"."+extension
                        try:
                                self[filetype+'path']=self[filetype+'path'].replace(fileserver_path,webfileserver)
                                # print(self[filetype+'path'])
                        except:
                                pass


    
    def get_components(self, components_only=True,qty=1):

        reflist=[]
        flatbom=[]

        def loopchildren(partnumber,revision,qty,reflist):
            part=mongoPart.objects(partnumber=partnumber,revision=revision)[0]
            children=part.getchildren()
            for child in children:
                refqty=child['qty']*qty
                if  len(child['part']['children'])>0:
                    if child['part'].hasConsumingProcess() and components_only:
                        reflist.append((child['part'],refqty))
                    else:
                        reflist.append((child['part'],refqty))
                        loopchildren(child['part']['partnumber'],child['part']['revision'],refqty,reflist)
                else:
                    reflist.append((child['part'],refqty))
                    
        loopchildren(self.partnumber,self.revision,qty,reflist)
        
        #Sum up all quantities and compile flatbom
        resdict={}
        for item,q in reflist:
            total=resdict.get(item,0)+q
            resdict[item]=total
        
        for part in resdict.keys():
            part.qty=resdict[part]
            flatbom.append(part)
        
        #Range flatbom by partnumber
        #flatbom.sort(key=lambda x: x.partnumber)
        print(flatbom)
        #input("stop")



        # flatbom.sort(key=lambda x: (x.category,x.supplier,x.approved,x.partnumber))
        
        totalqty=0
        for item in flatbom:
            totalqty=totalqty + item.qty
        
        print("Total components ", totalqty)        
        print("Unique components", len(flatbom))        
        return flatbom


    





def file_exists(location):

    if "http" in location:
        request = urllib.request.Request(location)
        request.get_method = lambda : 'HEAD'
        try:
            response = urllib.request.urlopen(request)
            return True
        except urllib.error.HTTPError:
            return False
    else:
        if os.path.isfile(location):
            return True
        else:
            return False


def web_to_pdf(url,fileout):
    config= pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    options = {
                'quiet': '' 
                }
    try:
        pdfkit.from_url(url, fileout, options=options)
    except:
        print("Couldn't export to pdf ",url)


#To find the encoding of a particular file
def find_encoding(fname):
    r_file = open(fname, 'rb').read()
    result = chardet.detect(r_file)
    charenc = result['encoding']
    return charenc

def create_folder_ifnotexists(path):    
#Check if outputfolder exists otherwise create it
        foldercheck=os.path.isdir(path)
        if not foldercheck:
            os.makedirs(path)



#To create thumbnails of images
def thumbnail(infile, size=(100, 100)):

    outfile = os.path.splitext(infile)[0] + ".thumbnail.png"
    if file_exists(outfile):
        if os.path.getatime(infile)>os.path.getatime(outfile):
            try:
                os.remove(outfile)
                im = Image.open(infile)
                im.thumbnail(size, Image.ANTIALIAS)
                im.save(outfile, "PNG")
                #print(outfile)
                return outfile
            except:
                print("Couldnt update existing OLD thumbnail - ",outfile)
                return outfile
        else:
            #print(outfile)
            return(outfile)

    else:
        try:
            im = Image.open(infile)
            im.thumbnail(size, Image.ANTIALIAS)
            im.save(outfile, "PNG")
            #print(outfile)
            return outfile
        except IOError:
            print ("cannot create thumbnail for '%s'" % infile)
            return ""



#To create a QR code to point to part link in tiny:
def qr_code(part,persist=True):
    # flash(part.partnumber)

    qrfile=fileserver_path+"/Deliverables/png/"+part.partnumber+"_REV_"+part.revision+".qr.jpg"


    if file_exists(qrfile):
        try:
            os.remove(qrfile)
            # flash("erased"+ qrfile)
            
        except:
            # flash("couldnt earse" + qrfile)
            pass


    image_url="http://"+webserver+"/part/"
    image_url+= part.partnumber+"_rev_"

    if part.revision=="":
                image_url+= "%25"
    else:
                image_url+=part.revision

    image_url=image_url.replace(" ","%20")

    #Creating an instance of qrcode

    qr = qrcode.QRCode(
            version=1,
            box_size=10,
            border=5)
    qr.add_data(image_url)
    qr.make(fit=True)

    img = qr.make_image(fill='black', back_color='white')
    tempfile=qrfile.replace('.qr.jpg','.temp.jpg')
    tempfile=img.save(qrfile)
    img.close()

    if persist:
        part['qrpath']=qrfile
        part.save()

    return qrfile








    


#To find the children based on a flatbom and a bom per the class definition below of solidbom
def get_children(father_partnumber,father_rev,bom,flatbom, qty="total"):
        children=bom.loc[(bom['father_partnumber']==father_partnumber) & (bom['father_revision']==father_rev)]
        children_rename_dict={}; children_rename_dict['child_partnumber']='partnumber'
        children_rename_dict['child_revision']='revision'
        children=children.rename(columns=children_rename_dict)

        
        children_flatbom=flatbom.merge(children,on=['partnumber','revision'],
                          how='left',indicator=True).query('_merge == "both"').drop(columns='_merge').reset_index(drop=True).sort_values(by='partnumber')

        if len(children)>0:
            return children_flatbom
        else:
            return []
    
        




# class User(db.Model):
#     # Defines the Table comment
#     __tablename__ = "user"

#     id = db.Column(db.Integer, primary_key=True, autoincrement=True)
#     username = db.Column(db.String,unique=True , nullable=False)
#     password = db.Column(db.String, nullable=False)
#     email = db.Column(db.String, nullable=False)
#     password = db.Column(db.String)
#     authenticated = db.Column(db.Boolean, default=False)

#     def is_active(self):
#         """True, as all users are active."""
#         return True

#     def get_id(self):
#         """Return the email address to satisfy Flask-Login's requirements."""
#         return self.email

#     def is_authenticated(self):
#         """Return True if the user is authenticated."""
#         return self.authenticated

#     def is_anonymous(self):
#         """False, as anonymous users aren't supported."""
#         return False


# class Comment(db.Model):
#     # Defines the Table comment
#     __tablename__ = "comment"
    
#     id = db.Column(db.Integer, primary_key=True, autoincrement=True)
#     part_id = db.Column(db.Integer, ForeignKey('part.id'), nullable=False)
#     user_id = db.Column(db.Integer,ForeignKey('user.id'), nullable=False)
#     body =db.Column(db.String)
#     category =db.Column(db.String)
#     created=db.Column(db.Date)
#     pic_path=db.Column(db.String)



    
#     def __init__(self, part_id="",user_id="",body="",category="", created="",pic_path=""):
#         self.part_id = part_id
#         self.user_id = user_id
#         self.body=body
#         self.category =category
#         self.pic_path=pic_path
#         self.created=created




###########################################################################################################
###########################################################################################################
###########################################################################################################
###########################################################################################################


class mongoJob(DynamicDocument):
    meta = {'collection': 'job'}

    jobnumber=StringField(unique=True )
    description=StringField( )
    customer=StringField( )
    user_id=StringField( )
    date_create=StringField( )
    date_due=StringField( )
    date_modify=StringField( )
    date_finish=StringField( )
    bom=ListField(DynamicField())



    #For the print outs on terminal of the object
    def __repr__(self):
        return f'Job/n_{self.jobnumber}_DES_{self.description}'
    def __str__ (self):
        return f'Job/n_{self.jobnumber}_DES_{self.description}'

    def to_dict(self):        
        dirtydict=self.to_mongo().to_dict()
        try:
            dirtydict['_id']=str(dirtydict['_id'])
        except:
            pass     
        return dirtydict



class mongoSupplier (DynamicDocument):
    meta = {'collection': 'supplier'}
    
    suppliername=StringField(unique=True )
    description=StringField( )
    location=StringField()
    address= StringField()
    processes= ListField(StringField())
    contact= StringField()
    

    #For the print outs on terminal of the object
    def __repr__(self):
        return f'Supplier_{self.suppliername}_DES_{self.description}'
    def __str__ (self):
        return f'Supplier_{self.suppliername}_DES_{self.description}'

    def to_dict(self):        
        dirtydict=self.to_mongo().to_dict()
        try:
            dirtydict['_id']=str(dirtydict['_id'])
        except:
            pass     
        return dirtydict

    

class mongoOrder (DynamicDocument):
    meta = {'collection': 'order'}

    ordernumber=StringField(unique=True )
    description=StringField( )
    job=StringField()
    supplier=ReferenceField(mongoSupplier)
    parts=ListField(ReferenceField(mongoPart))
    user_id=StringField( )
    date_create=StringField( )
    date_due=StringField( )
    date_modify=StringField( )
    date_finish=StringField( )

    #For the print outs on terminal of the object
    def __repr__(self):
        return f'Job/n_{self.ordernumber}_DES_{self.description}'
    def __str__ (self):
        return f'Job/n_{self.ordernumber}_DES_{self.description}'

    def to_dict(self):        
        dirtydict=self.to_mongo().to_dict()
        try:
            dirtydict['_id']=str(dirtydict['_id'])
        except:
            pass     
        return dirtydict
















class solidbom():
    
    def __init__(self, bomfile, flatfile,deliverableslocation,outputfolder,toppart=None):
            
            ###Load configuration
            [self.process_conf,self.property_conf, self.variables_conf] = [process_conf , property_conf,variables_conf]

        
            ### List of invalid chars, names, etc...
            ### this is to be set up from a conf file or something
            self.file=bomfile
            self.invalid_col_chars=['.','-',' ']
            self.renamelist={'part number':'partnumber',
                        'sw-configuration name(configuration name)':'sw_configuration',
                        'sw-folder name(folder name)':'folder',
                        'sw-file name(file name)':'file',
                        'item no.':'item_no'}
            #to rename teh cols
            self.renamedict={}
            
            #to drop the cols that are not required
            self.dropcols=[]
            
            self.col_clean_list=[]
            
            #Variable types
            self.int_cols=['qty'] 
            # self.float_cols=['mass','thickness']
            # self.bool_cols=['spare_part']
            
            #Add timestamp of bom creation
            self.timestamp=date.today().strftime('%d_%m_%Y')
            
                       
            #Default folder out
            self.folderout=outputfolder
            self.deliverables_folder=deliverableslocation
            
            
            
            
            ##### Load the input file and create the dataframe
            if bomfile!="" and flatfile!="":
               
                print(flatfile)

                #Insert all the individual part files, it will override the properties of the previous ones
                with open(flatfile) as f:
                    lines=f.readlines()
                    print(lines)
                

                for line  in lines:  
                    print(line)
                    partdict=json.loads(line)
                    print(partdict)


                    #Add uploader user
                    partdict['uploader']=current_user.to_dict()
                    

                    partnumber=partdict["partnumber"]
                    revision=partdict["revision"]

                    print("IMPORTED DICT - ",partnumber,revision)
                    existing=partcol.find_one({"partnumber":partnumber,"revision":revision})

                    print(existing)
                    # flash(existing)

                    if existing==None:
                        partcol.insert_one(partdict)
                        # print(input("test"))

                    else:
                        
                        fieldsdrop={}

                        for field in existing.keys():
                            if field != '_id' and field!='partnumber' and field!='revision':
                                fieldsdrop[field]=""
                            
                        # partcol.update ( { "_id": mongoid },{ "$unset":fieldsdrop})
                        # partcol.update ( { "_id": mongoid },{ "$set":partdict})

                        partcol.update_one({"partnumber":partnumber, "revision":revision},{ "$unset":fieldsdrop})
                        partcol.update_one({"partnumber":partnumber, "revision":revision},{ "$set":partdict})
                        print("-----------------------------------------------------")
                        print("-----------------------------------------------------")
                        print(partdict)
                        print("-----------------------------------------------------")
                        print("-----------------------------------------------------")
                    
                    part=mongoPart.objects(partnumber=partnumber,revision=revision)[0]
                    part.updateFileset(persist=True)
                    if part.pngpath:
                        cropandbackground(part.pngpath)
                    
                    part.get_process_icons(persist=True)


                    #To add the coating process if specified
                    if "zinc" in part.finish.lower():
                        part.process.append("zinc")
                        part.get_process_icons(persist=True)
                    
                    if "gal" in part.finish.lower():
                        part.process.append("galvanize")
                        part.get_process_icons(persist=True)
                    
                    if "nickel" in part.finish.lower():
                        part.process.append("nickel")
                        part.get_process_icons(persist=True)

                    



                    qr_code(part, persist=True)

                    #if "pngpath" in part.keys():



 

                ####BOM IMPORT STARTS ######
                file_enc=find_encoding(bomfile)
                self.filedata=pd.read_csv(bomfile, encoding =file_enc, sep='\t', lineterminator='\r',dtype=str)
                
                #Copy data to maniulate
                self.data=self.filedata.fillna("").copy()
                print("************** Afterimport filling nan with "" ")
                print(self.data)
                
                #Rename cols, not dependent on excel config file, check for future
                self.data=self.data.rename(columns={"ITEM NO.":"item_no","PART NUMBER":"partnumber","Approved":"approved","QTY.":"qty"}) 

                #Remove leading spaces from all cols
                for col in self.data:                    
                    if col!='qty': self.data[col]=self.data[col].str.lstrip()
                print("************** After removing spacess")
                print(self.data)
                                
                #Drop rows with no quantity and transform the col to int
                self.data.dropna(subset = ["qty"], inplace=True)
                self.data["qty"]=pd.to_numeric(self.data["qty"], errors='coerce').fillna(0).astype(int)
    
                print("************** after as rename, drop no quantities, and to int")
                print(self.data)

                #Drop empty partnumbers rows
                self.data=self.data[self.data.partnumber!=""]
                print("************** after dropping no partnumber rows ")
                print(self.data)


                #Modify the revision entry to account for the approved status
                #self.data.loc[self.data['event'].eq('add_rd') & self.data['environment']=="", 'environment'] = 'RD'

                self.data['revision']=np.where(self.data['approved']=="",self.data['revision']+"MOD",self.data['revision'])
                self.data.reset_index(drop=True)
                print("************** modifiyng revision based on approved ")
                print(self.data)


                
                ################################################
                ######################################################
                #Apply company customization 
                #no need with mongodb
                # self.customize()
                
                #Clean the entry data
                # self.clean_data()
                
                #Get the root component definition
                print(self.data)
                self.root_definition()
                
                
                            
                #Create bom objects
                self.createbom()
                
                #Create bom objects
                self.uploadbom()

                print("**********************************")
                print("**********************************")
                
                #Screening based on custom properties
                # self.property_screening()
                
                #Find related files/deliverables
                # self.find_deliverables()
            
                #Dataframe to database
                #self.createdatabase()
                
            elif toppart!=None:
                
                self.part=toppart
                self.partnumber=toppart.partnumber
                self.revision=toppart.revision
                self.description=toppart.description
                #self.flatbom=get_flatbom(self.partnumber,self.revision)
                self.tag=self.partnumber+"-REV-"+self.revision+"-"+self.timestamp


    def solidbom_from_flatbom(object_list,part,outputfolder="",sort=False):
        if outputfolder=="":
            folderout="temp"
        else:
            folderout="outputfolder"
        print("---obj in--- for " , part.partnumber)
        for pepe in object_list:
            print(pepe.partnumber)
        bomout=solidbom ( "","",deliverables_folder,folderout+"/" , toppart=part)

        bomout.flatbom=pd.DataFrame([x.to_dict() for x in object_list])

        print("---bom out --")
        for i,row in bomout.flatbom.iterrows():
            print(row.partnumber)
        return bomout
        print("---end --")                
        print(" ")                
            
    # def find_deliverables(self):

    #     #Create the empty cols for storing the info
    #     for extension in deliverables:
    #         self.flatbom[extension]=False
    #         self.flatbom[extension+'path']=''
        
        
    #     for i in range(len(self.flatbom)):
    #         partstring=self.flatbom.at[i,'file']+"_REV_"+self.flatbom.at[i,'revision']
            
    #         for extension in deliverables:
    #             extension_folder=self.deliverables_folder+extension+"/"
    #             file_string=extension_folder+partstring+"."+extension
    #             if os.path.isfile(file_string): 
    #                 self.flatbom.at[i,extension+'path']=file_string
    #                 self.flatbom.at[i,extension]=True
            
       
             
    def root_definition(self):
        self.partnumber=self.data.at[0,'partnumber']
        self.revision=self.data.at[0,'revision']
                
        self.tag=self.partnumber+"-REV-"+self.revision+"-"+self.timestamp
        
        #Set output location
        self.folderout=self.folderout+self.tag+"/"
        
        create_folder_ifnotexists(self.folderout)
        
        
    # #no need with mongodb             
    # def customize(self):
        
           
    #     #Build the rename dictionary and rename cols
    #     # for prop in property_conf.keys():
    #     #     self.renamedict[property_conf[prop]['custom_property']]=prop
        
    #     # self.data=self.data.rename(columns=self.renamedict) 
        
    #     #Drop the non required properties from the dataframe
    #     # for col in self.data.columns:
    #     #     if not col in [*property_conf]: 
    #     #         self.dropcols.append(col)
    #     # self.data= self.data.drop(self.dropcols, axis = 1)

    #     #Rename cols, see above to use the input from the excel file later on
    #     self.data=self.data.rename(columns={"ITEM NO.":"item_no","PART NUMBER":"partnumber","Approved":"approved","QTY.":"qty"}) 
        
       
        
      
             
           
    def clean_data(self):
        
        ###Remove all the entries with no file related to it
        ###This will hasve to be revised if the cut list and sheet metal properties
        ### are to be counted
#        self.data.dropna(subset = ["qty"], inplace=True)


        #Transform to right data type based on init lists:
        #Nan values will be filled with empty string or 0

        #no need with mongodb
        for col in self.data:
            #Remove spaces
            self.data[col]=self.data[col].str.lstrip()
            
            if col in self.int_cols:
                try:
                    self.data[col]=pd.to_numeric(self.data[col].fillna(0).astype(int))
                except:
                    print("Problems tranforming to int " , col)
                    print("forcing and replacing non valid with 0.0, check impact on output files")
                    self.data[col]=pd.to_numeric(self.data[col], errors='coerce').fillna(0).astype(int)
                
        #     elif col in self.float_cols:
        #         try:
        #             self.data[col]=pd.to_numeric(self.data[col].fillna(0.0).astype(float))
        #         except:
        #             print("Problems tranforming to float " , col)
        #             print("forcing and replacing non valid with 0.0, check impact on output files")
        #             self.data[col]=pd.to_numeric(self.data[col], errors='coerce').fillna(0.0).astype(float)
        #     elif col in self.bool_cols:
        #         self.data[col]=self.data[col].fillna(False).astype(bool)
        #     else:
        #         self.data[col]=self.data[col].fillna("").astype(str)
            
        #no need with mongodb
        # #Remove empty rows, spaces and and dirty data
        # self.data['revision']=self.data['revision'].astype(str)
        
        #Remove non alfanumeric from revision and replace by "" (in case empty or funny)
        self.data.loc[self.data.revision.str.isalnum()==False,'revision']=''
        self.data.loc[self.data.revision=="nan",'revision']=''
        self.data.loc[self.data.revision.str.isalnum()==False,'approved']=''
        self.data.loc[self.data.revision=="nan",'approved']=''
        
        
        #Remove leading spaces in part number
        print(self.data)
        self.data['partnumber']=self.data['partnumber'].str.lstrip()

        #Remove all the empty partnumber lines (to include in the future )
        self.data.dropna(subset = ["partnumber"], inplace=True)

        print(self.data)
        
        
        #no need with mongodb
        #Reorder the columns for easier debugging and manipulations
        # newcols=maincols+refcols
        # for col in self.data.columns.to_list():
        #     if not col in newcols: newcols.append(col)
        
        # self.data=self.data.reindex(columns=newcols)
        
        #no need with mongodb
        #Put process, finish and treatment as lowercase
        # for prop in lowercase_properties:
        #     self.data[prop]=self.data[prop].str.lower()
      
        #Remove all non desired properties
        

         
        

    def property_screening(self):

              
        #Account for the hardware in process related
        for folder in hardware_folder:
            self.flatbom.loc[self.flatbom['folder'].str.lower().str.contains(folder),['process','process2','process3']]=['hardware',"",""]
       
        
       
  
        
    def createdatabase(self):
        
        
        # #Add parts to database
        # #The overwrite must be revised
        # for index, row in self.flatbom.iterrows():
        #     database_part=Part()
        #     database_part.partfromlist(row)

        #     #Find if the part is already in database
        #     check_part=db.session.query(Part).filter(and_(Part.partnumber==row['partnumber'] ,
        #                                              Part.revision==row['revision'])
        #                                              ).first()
        #     #Create or update qrcode
        #     qr_code(database_part)

        #     #IF part exists overwrite attributes
        #     if check_part==None:
        #         db.session.add(database_part)
                

        #     else:
        #         for prop in property_conf.keys():
        #             if hasattr(database_part, prop):
        #                 setattr(check_part,prop,getattr(database_part,prop))
                        
      
        # #Commit changes on part table
        # db.session.commit() 
        # db.session.close()
        


        #Add bom to database
        for index, row in self.flatbom.iterrows():            
        
            database_part=Part()
            database_part=db.session.query(Part).filter(and_(Part.partnumber==row['partnumber'],
                                                     Part.revision==row['revision'])
                                                     ).first()

                        
            #Erase all bom entries of related part downstrema, erase kids entries
            if database_part.partnumber!="":
                bomentries=db.session.query(Bom).filter(Bom.father_id==database_part.id)
                for bomline in bomentries:
                    db.session.delete(bomline)
                db.session.commit() 

                #Get children from solidbom
                children=get_children(database_part.partnumber,database_part.revision,self.bom, self.flatbom)

                #Add children to database
                if len(children)>0:
                    for i, childrow in children.iterrows():
                        childpart=db.session.query(Part).filter(and_(Part.partnumber==childrow['partnumber'] ,
                                                     Part.revision==childrow['revision'])
                                                     ).first()
                        if childpart!=None:
                            bomentry=Bom(database_part.id,childpart.id,childrow['qty'])
                            db.session.add(bomentry)

        #Commit changes on bom table
        db.session.commit() 
        db.session.close()


        
    def createbom(self):

        self.flatbom=self.data.copy()

        #Remove qty and tree reference to  get unique entries
        #And add the totalquanty
        del self.flatbom['item_no']
        del self.flatbom['qty']
        self.flatbom['totalqty']=0
            
        # #Drop flatbom duplicates
        self.flatbom= self.flatbom.drop_duplicates()
        self.flatbom=self.flatbom.reset_index(drop=True)
        
        
        ## Create bom dataframe
        self.bom= pd.DataFrame({'father_partnumber':"",   
                'father_revision':"",
                'child_partnumber':self.data['partnumber'],
                'child_revision':self.data['revision'],
                'qty':self.data['qty'],
                'ref': self.data ['item_no']
                })
        
        #self.bom.reset_index(drop=True)

        #Filter bom ref for pointing only to father for 
        #accouting of duplicated configurations
        for index, row in self.data.iterrows():
            self.bom.at[index,'ref']=re.sub(r"\..?[0-9]$", "",self.bom.at[index,'ref'])
        
        
        #Build bom table finding referenced father part number
        for index, row in self.data.iterrows():
                
                temp=self.data.loc[self.data['item_no'] ==re.sub(r"\..?[0-9]$", "", row ['item_no'])].reset_index(drop=True)
                self.bom.at[index,'father_partnumber']=temp.at[0,'partnumber']
                self.bom.at[index,'father_revision']=temp.at[0,'revision']
                #If couldnt repalce, it means twe are in the top level and we add the mark root
                if self.bom.at[index,'father_partnumber']==self.bom.at[index,'child_partnumber']:
                    self.bom.at[index,'father_partnumber']='root'
  
       #Combine duplicate entries quantities and reset index
        self.bom=self.bom.groupby(['child_partnumber','child_revision','father_partnumber','father_revision','ref'])['qty'].sum().to_frame().reset_index()
        self.bom.reset_index(drop=True)
        
        
      
         # #Compute the total quantity for each partnumber
        self.data['branchqty']=self.data['qty']
        self.data=self.data.set_index('item_no')

    
        for index, row in self.data.iterrows():
            father_index=re.sub(r"\..?[0-9]$", "", index)
            while len(father_index.split('.'))>1:
                self.data.at[index,'branchqty']=self.data.at[index,'branchqty']*self.data.at[father_index,'branchqty']
                father_index=re.sub(r"\..?[0-9]$", "", father_index)
            
        for index, row in self.flatbom.iterrows():
            
            self.flatbom.at[index,'totalqty']=self.data[(self.data['partnumber']==row['partnumber'] )& (self.data['revision']==row['revision'])]['branchqty'].sum()


        print("************** BOM ")
        print (self.bom)
        print("************** FLATBOM ")
        print (self.flatbom)
        print("************** DATA AFTER bOM ")
        print(self.data)


    def uploadbom(self):

        for index, row in self.bom.iterrows():

            fatherPN=row['father_partnumber']
            childPN=row['child_partnumber']
            fatherREV=row['father_revision']
            childREV=row['child_revision']
            qty=row['qty']

            father=partcol.find_one({"partnumber":fatherPN,"revision":fatherREV})   
            child=partcol.find_one({"partnumber":childPN,"revision":childREV})

            # father=partcol.find_one({"partnumber":fatherPN})   
            # child=partcol.find_one({"partnumber":childPN})

            # print("*"+fatherPN+"*","*"+ fatherREV+"*")
            # print("*"+childPN+"*","*"+ childREV+"*")

            # print(father['partnumber'], child['partnumber'])
            # print(father['partnumber'], child['partnumber'])
            
            if father!=None and child!=None:
                if "children" in father.keys():
                    partcol.update_one({"partnumber":fatherPN, "revision":fatherREV},{ "$push":{'children': child['_id'],'childrenqty': qty}})
                else:
                    partcol.update_one({"partnumber":fatherPN, "revision":fatherREV},{ "$set":{'children':[ child['_id']],'childrenqty':[qty]}})
            
        

        
    def solidbom_to_excel(self,process=""):
        redFill = PatternFill(start_color='FFEE1111',
                      end_color='FFEE1111',
                      fill_type='solid')
        yellowFill = PatternFill(start_color='00FFFF00',
                              end_color='00FFFF00',
                              fill_type='solid')
        
        bom_in=self.flatbom.loc[(self.flatbom['process']!='hardware') & (self.flatbom['process2']!='hardware') & (self.flatbom['process3']!='hardware')]
       

        #Rearrange the cols
        notmain_cols=[]
        last_cols=[]
        
        for col in bom_in:
            if col in maincols or col in refcols:
                pass
            elif "path" in col or col in deliverables or col in ['eprt','edrw','easm','threemf']:
                last_cols.append(col)
            elif col!="_sa_instance_state":
                notmain_cols.append(col)

        #Add screenshot col
        bom_in["Screenshot"]=""

        #Copy of bom for bom generation only (avoid having images path collumns)
        bom_image=bom_in[['partnumber']+['revision']+['pngpath']+['png']]

        #Replace datasheet location by web link:
        bom_in['datasheet']=bom_in['datasheet'].str.replace("//","/")
        bom_in['datasheet']=bom_in['datasheet'].str.replace(fileserver_path,"http://"+webfileserver)
        




        if process!="":
            bom_in=bom_in[["Screenshot"]+process_conf[process]['fields']]
            
        else:
            bom_in=bom_in[["Screenshot"]+maincols+refcols+notmain_cols+last_cols]
    



        self.excel_file=self.folderout + "BOM_tables-" +self.tag  +".xlsx"

        #Prepare file and format wrap
        writer = pd.ExcelWriter(self.excel_file, engine = 'xlsxwriter')
        workbook = writer.book
        wrap_format = workbook.add_format({'text_wrap': True})
        #wrap_format.set_border(6)
        
        #Dump dataframe to excel
        if process!="":
            sheet_name = process.upper()+ ' scope of supply'
        else:
            sheet_name = 'Flatbom'
        
        bom_in.to_excel(writer, sheet_name = sheet_name)

         # Set the worksheet and  autofilter.
        worksheet = writer.sheets[sheet_name]
        (max_row, max_col) = bom_in.shape
        worksheet.autofilter(0, 0, max_row, max_col )
        worksheet.set_column(2,max_col,None,wrap_format)

        

        
        #Width of cols
        worksheet.set_column(0,0,4,wrap_format)
        worksheet.set_column(1,1,8,wrap_format)
        worksheet.set_column(2,2,20,wrap_format)
        worksheet.set_column(3,3,4,wrap_format)
        worksheet.set_column(4,4,30,wrap_format)
        worksheet.set_column(5,5,4,wrap_format)
        worksheet.set_column(6,max_col,20,wrap_format)




         # Add images section
         
        i=-1           
        for index, row in bom_image.iterrows():
            i=i+1
            thumb=thumbnail(row['pngpath'])
            print(thumb) 

            #Adjust row height
            worksheet.set_row(i+1,30)

            #Add image
            cell='B'+str(i+2)

            image_url="http://"+webserver+"/part/"
            image_url+= row['partnumber']+"_rev_"

            if row['revision']=="":
                image_url+= "%25"
            else:
                image_url+= row['revision']

            print(row['png'])

                

            if (row['png']!="FALSE" and  row['png']!=False ) or thumb!="":
                worksheet.insert_image(cell, thumb, {'x_offset': 1,
                                                    'y_offset': 1,
                                                    'x_scale': 0.5,
                                                    'y_scale': 0.5,
                                                    'object_position': 1,
                                                     'url': image_url})
        
        if process=="":
            ##Generate basic lists:
            #with pd.ExcelWriter(self.excel_file) as writer:
            listofprocess=[*process_conf]
            for process in listofprocess:
                property_list=['partnumber','revision','description','material']
                values_list=["totalqty"]
                    
                process_df=self.flatbom.loc[(self.flatbom['process']==process) | (self.flatbom['process2']==process) | (self.flatbom['process3']==process)]
                    
                if process=='folding' or process=='lasercut' or process=='profile cut' :
                    property_list.append('thickness')
                    
                if process=='purchase':
                    property_list=['partnumber','description','oem','supplier','supplier_partnumber', 'category','datasheet','pdfpath']
                if process=='hardware':
                    property_list.remove('revision')


                    
                process_df=pd.pivot_table(process_df,index=property_list,values=values_list)
                if len(process_df)>0: 
                        process_df.to_excel(writer,sheet_name=process)
                        #print (process_df)   
                        # Set the autofilter.
                        worksheet = writer.sheets[process]
                        (max_row, max_col) = process_df.shape
                        worksheet.autofilter(0, 0, max_row, max_col +len (property_list)-1)
                        worksheet.set_column(0,max_col +len (property_list)-1,15,wrap_format)

            

                
        #Close workbook with xlsxwriter so openpyxl can open it
        workbook.close()

        return self.excel_file
        
        

        
    def get_parents(self,partnumber,revision):
        parents=self.bom.loc[(self.bom['child_partnumber']==partnumber) & (self.bom['child_revision']==revision)]
        parents_rename_dict={} ; parents_rename_dict['father_partnumber']='partnumber'
        parents_rename_dict['father_revision']='revision'
        parents=parents.rename(columns=parents_rename_dict)
        
        parents=self.flatbom.merge(parents,on=['partnumber','revision'],
                          how='left',indicator=True).query('_merge == "both"').drop(columns='_merge').reset_index(drop=True)
        
        if len(parents)>0:
            return parents
        else:
            return []
        
        
    def gather_datasheet(self):
        
        #Check if outputfolder exists otherwise create it
        outputfolder=self.folderout+"datasheets/"
        create_folder_ifnotexists(outputfolder)
        
        if len(self.flatbom[(self.flatbom['process']=='purchase') |( self.flatbom['process2']=='purchase')|( self.flatbom['process3']=='purchase')]):
            purchasefolder=self.folderout+"purchase/"
            create_folder_ifnotexists(purchasefolder)
            
        for i in range(len(self.flatbom)):
            
            targetfile=outputfolder + self.flatbom['partnumber'][i]+"-datasheet"
    
            if os.path.isfile(self.flatbom.at[i,"datasheet"]):
                sourcefile=self.flatbom["datasheet"][i]
                fileName, fileExtension = os.path.splitext(sourcefile)
                

                
                if "png" in fileExtension.lower() or "jpg" in fileExtension.lower() : 
                    try:
                        image1=Image.open(sourcefile)
                        im1=image1.convert('RGB')
                        targetfile=targetfile+".pdf"
                        im1.save(targetfile)
                        self.flatbom.at[i,'datasheet']=targetfile
                    except:
                        print (self.flatbom['partnumber'][i] + " PROBLEMS COMPILING DATASHEET ", sourcefile)
                elif "pdf" in fileExtension.lower():    
                    targetfile=targetfile + fileExtension
                    sourcefile=PureWindowsPath(sourcefile)
                    targetfile=PureWindowsPath(targetfile)
                    #print(sourcefile, " copy to ", targetfile)
                    try:
                        copyfile(sourcefile,targetfile)
                        
                        self.flatbom.at[i,'datasheet']=targetfile
                    except:
                        print (self.flatbom['partnumber'][i] + " PROBLEMS COMPILING DATASHEET ", sourcefile, targetfile)
                        
                if self.flatbom['process'][i]=='purchase' or self.flatbom['process2'][i]=='purchase' or self.flatbom['process3'][i]=='purchase' :
                        purchasefile=purchasefolder + self.flatbom['partnumber'][i]+".pdf"
                        try:
                            copyfile(targetfile,purchasefile)
                        except:
                            print("Couldn't copy ", purchasefile  , " to ", targetfile)
                    
            elif self.flatbom.at[i,"datasheet"]!="":
                    try:
                        web_to_pdf(self.flatbom['link'][i],targetfile+".pdf")
                        print(self.flatbom['partnumber'][i] , " DATASHEET FROM WEB")
                        
                    except:
                        try:
                            print(self.flatbom['partnumber'][i] , " NO DATASHEET - ", sourcefile)
                        except:
                            print(self.flatbom['partnumber'][i] , " invalid SOURCEFILE !!!!!!!!!!")
                        # self.flatbom.at[i,'notes'].append("Invalid datasheet source")
    
    def gather_deliverables(self):
        
        #Loops over the processes in the excel configurator file 
        #but excludes the last one (others) that is why [*process_conf][:-1] 
        # so be careful when adding more
        
        for process in [*process_conf][:-1]:
            bom_in=self.flatbom.loc[(self.flatbom['process']==process) | (self.flatbom['process2']==process) | (self.flatbom['process3']==process)]
            
            if len(bom_in)==0:
                continue
            
            musthave=[]
            if process_conf[process]['pdf']==1: musthave.append("pdf")
            if process_conf[process]['dxf']==1: musthave.append("dxf")
            if process_conf[process]['step']==1: musthave.append("step")
            
            #Check if outputfolder exists otherwise create it
            outputfolder=self.folderout+process+"/"
            
            print(outputfolder)
            if len(musthave)>0:create_folder_ifnotexists(outputfolder)
    
            
            for index, row in bom_in.iterrows():
                #print(index,row['partnumber'], row["pdfpath"])  
                
                filenamebit=row["partnumber"]+"_REV_"+row["revision"]
                for extension in musthave:
                    sourcefile=self.deliverables_folder+extension+"/"+filenamebit+"."+extension
                    
                    if process=='folding' or process=='lasercut' or process=='profile cut':
                        targetfile=outputfolder+filenamebit
                        targetfile=targetfile+"-"+row["material"]+"_"+str(row["thickness"])+"mm."+extension
                        # if row["thickness"]==0 or str(row["thickness"])=='':
                           # self.flatbom.at[index,'notes'].append("Missing thickness")
                        
                    else:
                        targetfile=outputfolder+filenamebit+"."+extension
                    
                                        
                    if os.path.isfile(sourcefile):
                        copyfile(sourcefile,targetfile)
                print(sourcefile,targetfile)

  
class Bom(db.Model):
    # Defines the Table Name user
    __tablename__ = "bom"
    
	# Defines the variables
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    father_id = db.Column(Integer, ForeignKey('part.id'), nullable=False)
    child_id = db.Column(Integer, ForeignKey('part.id') , nullable=False)
    qty=db.Column(Integer, nullable=False)
        
    
    def __init__(self, father_id, child_id,qty):
        self.father_id = father_id
        self.child_id = child_id
        self.qty = qty
        self.child=Part.query.filter_by(id=self.child_id).first()
        
    def getchild(self):
        self.child=Part.query.filter_by(id=self.child_id).first()
        return self.child
        
    def __repr__(self):
        
        self.getchild()
        return f'BOM( {self.child.partnumber}, {self.child.revision} , quantity {self.qty})'
    def __str__ (self):
        self.getchild()
        return f'BOM( {self.child.partnumber} , {self.child.revision} ,quantity {self.qty})'    



def deletepart(database_part,echo=False):
              
    bomentries=db.session.query(Bom).filter(or_(Bom.father_id==database_part.id,Bom.child_id==database_part.id))
    for bomline in bomentries:
        this=bomline.id
        db.session.delete(bomline)
        if echo: print("deleted-",this)
    db.session.commit()
    # bomentries=db.session.query(Bom).filter(Bom.child_id==database_part.id)
    # for bomline in bomentries:
    #     this=bomline.id
    #     db.session.delete(bomline)
    #     if echo: print("deleted-",this)
    # db.session.commit()
    this=database_part.id
    
    db.session.delete(database_part)
    if echo: print("part deleted-",this)

    db.session.commit()
    if echo: print("all deleted from ",this)
    #db.session.close()
 

class Part(db.Model):
    # Defines the Table Name user
    __tablename__ = "part"
    
	# Makes three columns into the table id, name, email
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    partnumber = db.Column(db.String, nullable=False)
    revision = db.Column(db.String)
    
    approved = db.Column(db.String)
    author = db.Column(db.String)
    category = db.Column(db.String)
    configuration = db.Column(db.String)
    colour = db.Column(db.String)
    datasheet = db.Column(db.String)
    description = db.Column(db.String)
    drawndate = db.Column(db.String)
    file = db.Column(db.String)
    finish = db.Column(db.String)
    folder = db.Column(db.String)
    link = db.Column(db.String)
    mass = db.Column(db.Float)
    material = db.Column(db.String)
    oem = db.Column(db.String)
    
    process = db.Column(db.String)
    process2 = db.Column(db.String)
    process3 = db.Column(db.String)
    
    spare_part = db.Column(db.Boolean)
    supplier = db.Column(db.String)
    supplier_partnumber = db.Column(db.String)
    thickness = db.Column(db.Float)
    treatment = db.Column(db.String)
    colour = db.Column(db.String)
    notes = db.Column(db.String)
    asset = db.Column(db.String)
    
    children= relationship("Part", 
                           secondary="bom",
                           primaryjoin=id==Bom.father_id,
                           secondaryjoin=id==Bom.child_id,
                           backref="parent" )
    # comments=relationship("Comment")

    def as_dict(self,folder=""):
        if not hasattr(self,'qty'):
            self.qty=0
            self.totalqty=self.qty
        else:
            self.totalqty=self.qty

        self.datasheet_available=False
        self.png=False
        self.pdf=False
        self.eprt=False
        self.edrw=False
        self.easm=False
        self.edr=False
        self.png_d=False
        self.dxf=False
        self.step=False
        self.threemf=False
        
        self.datasheet_link=""   
        self.modelpath=""
        self.pngpath=""
        self.pdfpath=""
        self.dxfpath=""
        self.edrpath=""
        self.steppath=""
        self.threemfpath=""
        self.png_dpath=""

        if folder!="":
            self.updatefilespath(folder,local=True)

        return self.__dict__

    def to_dict(self):
        self.updatefilespath(webfileserver)

        
        if self.revision=="":
            urllink=url_for('tinylib.partnumber',partnumber=self.partnumber,revision="%25",detail="quick")
        else:
            urllink=url_for('tinylib.partnumber',partnumber=self.partnumber,revision=self.revision,detail="quick")

        weblink= '<a href="'+ urllink +  '">' + """<img src="http://""" + self.pngpath + """" width=auto height=30rm></a>"""
        
     


        return {
            'id': self.id,
            'pngpath': weblink,
            'partnumber': self.partnumber,
            'revision': self.revision,
            'description': self.description,
            'process': self.process,
            'process2': self.process2,
            'process3': self.process3,
            'finish': self.finish
        }
        
    
    def getchildren (self):
        outlist=[]
        
        for child in self.children:
            kid={}
            kid['part']=child
            kid['qty']=Bom.query.filter_by(father_id=self.id,child_id=child.id).first().qty
            outlist.append(kid)
        return outlist
    
    def children_with_qty (self):
        outlist=[]
        
        for child in self.children:
            child.qty=Bom.query.filter_by(father_id=self.id,child_id=child.id).first().qty
            outlist.append(child)

        return outlist

    def get_tag(self):
        self.tag=self.partnumber+"_REV_"+self.revision+"-"+date.today().strftime('%d_%m_%Y')
        return self.tag


    def parents_with_qty (self):
        outlist=[]
        
        for parent in self.parent:
            parent.qty=Bom.query.filter_by(father_id=parent.id,child_id=self.id).first().qty
            outlist.append(parent)

        return outlist

    def hasProcess(self,process):
        if process in self.process or process in self.process2 or process in self.process3:
            return True
        else:
            return False
    
             



    def hasConsumingProcess(self):
        processlist=[self.process,self.process2,self.process3]
        
        processlist=[x for x in processlist if x in process_conf.keys()]

        consume=False

        for process in processlist:
            if int(process_conf[process]['priority'])<20:
                consume=True

        return consume



    def isMainProcess(self,process):
        mainprocess_bool=False

        if self.hasProcess(process):
            processlist=[self.process,self.process2,self.process3]
            processlist=[x for x in processlist if x in process_conf.keys()]
            if len(processlist)>0:
                for x in processlist:
                    if int(process_conf[process]['priority'])<= int(process_conf[x]['priority']):
                        mainprocess_bool=True
                        
                    else:
                        mainprocess_bool=False
            

        return mainprocess_bool


    def MainProcess(self):
        processlist=[self.process,self.process2,self.process3]
        processlist=[x for x in processlist if x in process_conf.keys()]

        mainprocess=self.process
        if len(processlist)>0:
            for x in processlist:

                if int(process_conf[x]['priority'])<= int(process_conf[mainprocess]['priority']):
                    mainprocess=x

        else:
            mainprocess='other'
        self.mainprocess=mainprocess
                
        return mainprocess



    def __repr__(self):
        return f'part( {self.partnumber} , {self.revision} , {self.description})'
    def __str__ (self):
        return f'part( {self.partnumber} , {self.revision} , {self.description})'

    def updatefilespath(self,folderin,local=False, png_thumbnail=False):
        #This functions check first in the file server local path and then adds the http link
        folder=folderin
        # folder=webfileserver

        self.tag=self.partnumber+"_REV_"+self.revision

        pngfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"png/"
        pdffolder=fileserver_path+variables_conf['deliverables_folder']['value']+"pdf/"
        edrfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"edr/"
        stepfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"step/"
        dxffolder=fileserver_path+variables_conf['deliverables_folder']['value']+"dxf/"
        threemffolder=fileserver_path+variables_conf['deliverables_folder']['value']+"3mf/"

        #Ssame for datasheets
        datasheetfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"/datasheet/"
        
        self.pdfpath=pdffolder+self.tag+".pdf"
        self.steppath=stepfolder+self.tag+".step"
        self.dxfpath=dxffolder+self.tag+".dxf"
        self.eprtpath=edrfolder+self.tag+".eprt"
        self.easmpath=edrfolder+self.tag+".easm"
        self.edrwpath=edrfolder+self.tag+".edrw"
        self.png_dpath=pngfolder+self.tag+"_DWG.png"
        self.pngpath=pngfolder+self.tag+".png"
        self.qrpath=pngfolder+self.tag+".qr.jpg"
        self.threemfpath=threemffolder+self.tag+".3mf"

        if self.hasProcess("hardware") or self.hasProcess("purchase") and not file_exists(self.pngpath):
            self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"
  
        #if self.hasProcess("hardware") and not file_exists(self.pngpath):
        #    if not local:
        #        self.pngpath=pngfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".png"
        #    else:
        #        self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"

        #Adding boolean value to false for all files
        self.datasheet_available=False
        self.png=False
        self.qr=False
        self.pdf=False
        self.eprt=False
        self.edrw=False
        self.easm=False
        self.edr=False
        self.png_d=False
        self.dxf=False
        self.step=False
        self.threemf=False
        
        
        #Add directly the link for datasheets withouth checking file exists
        
        if self.datasheet:
            #print(self.datasheet)
            
            self.datasheet.replace('"', '')
            self.datasheet.replace("file:///",'')
            self.datasheet.replace('%20',' ')
            # path, filename = os.path.split(self.datasheet)
            # flash(PureWindowsPath(self.datasheet).name)
            # flash(filename)
            filename=PureWindowsPath(self.datasheet).name
            self.datasheet=datasheetfolder+filename
            
            if file_exists(self.datasheet):
                self.datasheet_link=folder+variables_conf['deliverables_folder']['value']+"/datasheet/"+ filename
                self.datasheet_link=self.datasheet_link.replace(' ','%20')
                self.datasheet_available=True
            else:
                if file_exists(self.datasheet+".pdf"):
                     self.datasheet_link=folder+variables_conf['deliverables_folder']['value']+"/datasheet/"+ filename+".pdf"
                     self.datasheet_link=self.datasheet_link.replace(' ','%20')
                     self.datasheet_available=True
            #if hasattr(self,"datasheet_link"): 
            #    print(self.datasheet_link)
      
           

        #Model paths added directly
        self.modelpath=self.folder +self.file+ ".SLDPRT"
        if file_exists( self.modelpath.upper()+ ".SLDPRT"): 
            self.modelpath= self.modelpath+ ".SLDPRT" 
        if file_exists( self.modelpath+ ".SLDASM"): 
            self.modelpath= self.modelpath+ ".SLDASM" 


        #UPdate the folder location to the given in function, normally if not local
        #the folder path will be the fileserver address
        if not local:
            pngfolder=folder+variables_conf['deliverables_folder']['value']+"png/"
            pdffolder=folder+variables_conf['deliverables_folder']['value']+"pdf/"
            edrfolder=folder+variables_conf['deliverables_folder']['value']+"edr/"
            stepfolder=folder+variables_conf['deliverables_folder']['value']+"step/"
            dxffolder=folder+variables_conf['deliverables_folder']['value']+"dxf/"
            threemffolder=folder+variables_conf['deliverables_folder']['value']+"3mf/"

        #To reduce the image quality on the flatbom if needed (too much bandwidth)
        
        
        #Png path
        

        #print(self.pngpath)
        if file_exists(self.pngpath):
            # print("exist")
            if png_thumbnail:
                pngfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"png/"
                self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"
                self.pngpath=thumbnail(self.pngpath)
                
                try:
                    path, filename = os.path.split(self.pngpath)
                    
                    self.pngpath=folder+variables_conf['deliverables_folder']['value']+"png/"+filename
                    
                    self.png=True 
                    
                except:
                    print("Problems with ", self.partnumber)
                    print("pngfolder  ", fileserver_path)
                    print("folder  ", folder)
                    print("fileserver_path  ", fileserver_path)
                    
                    self.png=False 
            else:
                self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"
             

            if not local:
                self.pngpath=self.pngpath.replace(' ','%20')
  
        else:
            self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"  


            
            self.png=True 
 

        # #qr path
        
        # self.qrpath=qr_code(self)
        self.qrpath=pngfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".qr.jpg"
        self.qr=True 


    



        #png_d path
        if file_exists(self.png_dpath):
            self.png_dpath=pngfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+"_DWG.png"
            self.png_d=True 
    

        #pdf path
        if file_exists(self.pdfpath):
            self.pdfpath=pdffolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".pdf"
            self.pdf=True 


        #dxf path
        if file_exists(self.dxfpath):
            self.dxfpath=dxffolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".dxf"
            self.dxf=True 


        #step path
        if file_exists(self.steppath):
            self.steppath=stepfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".step"
            self.step=True 

        #threemf path
        if file_exists(self.threemfpath):
            self.threemfpath=threemffolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".threemf"
            self.threemf=True 

        #Model Edrawings path
        if file_exists(self.eprtpath):
            self.eprtpath=edrfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".eprt"
            self.eprt=True 
            self.edr=True
            self.edrpath=self.eprtpath
        elif file_exists(self.easmpath):
            self.easmpath=edrfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".easm"
            self.easm=True 
            self.edr=True
            self.edrpath=self.easmpath 

        #Drawing - Edrawings path
        if file_exists(self.edrwpath):
            self.edrwpath=edrfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".edrw"
            self.edrw=True
            self.edr_d=True
            self.edr_dpath=self.edrwpath

        self.get_process_icons()



    def get_process_icons (self):
        self.process_icons=[]
        self.process_colors=[]
        if self.process!="" and self.process in process_conf.keys() :
            self.process_icons.append('images/'+(process_conf[self.process]['icon']))
            self.process_colors.append(process_conf[self.process]['color'])
            if self.process2 in process_conf.keys() :
                self.process_icons.append('images/'+(process_conf[self.process2]['icon']))
                self.process_colors.append(process_conf[self.process2]['color'])
            if self.process3 in process_conf.keys() :
                self.process_icons.append('images/'+(process_conf[self.process3]['icon']))
                self.process_colors.append(process_conf[self.process3]['color'])
        else:
            self.process_icons.append('images/'+process_conf['others']['icon'])
            self.process_colors.append(process_conf['others']['color'])






    
    def __init__(self, partnumber="",revision="",description="",
                process="",process2="",process3="",
                finish="",path_to_model_file="",material="",matspec="",partType="",
                pdfpath="",edrpath="",edr_dpath="",jpgpath="",jpg_dpath="",pngpath="",png_dpath="",
                dxfpath="",dwgpath="",
                pdf="",edr="",edr_d="",jpg="",jpg_d="",png="",png_d="",dxf="",dwg="",check="",notes="",
                pdfindex="",pdfpages="",qty="",totalqty="",colour="",asset=""):

        self.partnumber=partnumber
        self.revision=revision
        self.description=description
        self.process=process
        self.process2=process2
        self.process3=process3
        self.finish=finish
        self.material=material
        self.colour=colour
        self.notes=notes
        self.asset=asset
       
      
        
        #outputfiles
        self.pdfpath=pdfpath
        self.pngpath=pngpath

                
        #Bom and index related
        self.qty=qty
        self.totalqty=totalqty
        self.pdfindex=pdfindex

        
    def ispartprocess(self,process):
       
        if process in self.process or process in self.process2 or process in self.process3:
            return True
        else:
            return False
   
 
    #To create a part from a dataframe row
    def partfromlist(self,datalist):
        self.spare_part=datalist['spare_part']
        self.mass=datalist['mass']
        self.thickness=datalist['thickness']
        
        self.category=datalist['category']
        self.datasheet=datalist['datasheet']
        
        self.partnumber=datalist['partnumber']
        self.process=datalist['process']
        self.process2=datalist['process2']
        self.process3=datalist['process3']
        self.configuration=datalist['configuration']
        self.file=datalist['file']
        self.folder=datalist['folder']
        self.supplier_partnumber=datalist['supplier_partnumber']
        self.description=datalist['description']
        self.finish=datalist['finish']
        self.material=datalist['material']
        self.revision=datalist['revision']
        self.approved=datalist['approved']
        self.author=datalist['author']
        self.supplier=datalist['supplier']
        self.link=datalist['link']
        self.oem=datalist['oem']
        self.treatment=datalist['treatment']
        self.drawndate=datalist['drawndate']
        self.colour=datalist['colour']
        #self.notes=datalist['notes']
        self.category=datalist['category']
        self.asset=datalist['asset']
                
        
        ##Extra values only needed for pdf list exports
        self.pdfpath=datalist['pdfpath']
        self.pngpath=datalist['pngpath']

        try:
            self.qty=str(int(datalist[ 'qty']))
        except:
            self.qty=0
        self.totalqty=str(int(datalist[ 'totalqty']))
        
        self.png=False
        if os.path.isfile(self.pngpath): self.png=True
   
        
        try:
            self.pdfindex=str(int(datalist[ 'pdfindex']))
        except:
            self.pdfindex=""
        
        
        return self
   
    
    def get_components(self, components_only=True):

        reflist=[]
        flatbom=[]

        def loopchildren(partnumber,revision,qty,reflist):
            part=Part.query.filter_by(partnumber=partnumber,revision=revision).first()
            children=part.children_with_qty()
            for child in children:
                refqty=child.qty*qty
                
                if  len(child.children)>0:
                    
                    if child.hasConsumingProcess() and components_only:
                        reflist.append((child,refqty))
                    else:
                        reflist.append((child,refqty))
                        loopchildren(child.partnumber,child.revision,refqty,reflist)
                    
                else:
                    reflist.append((child,refqty))
                    
        loopchildren(self.partnumber,self.revision,1,reflist)
        
        #Sum up all quantities and compile flatbom
        resdict={}
        for item,q in reflist:
            total=resdict.get(item,0)+q
            resdict[item]=total
        
        for part in resdict.keys():
            part.qty=resdict[part]
            flatbom.append(part)
        
        #Range flatbom by partnumber
        #flatbom.sort(key=lambda x: x.partnumber)
        flatbom.sort(key=lambda x: (x.category,x.supplier,x.oem,x.approved,x.partnumber))
        
        print(len(flatbom))        
        return flatbom


    def partlist(bom_list):
        
        part_list=[]
        
        for i in range (len(bom_list)):
            part_in=Part()
            part_in=part_in.partfromlist(bom_list.iloc[i])
            part_list.append(part_in)
        return part_list
    

def get_tree(partnumber,revision,partlist,qty=1):
    
    refpart=Part.query.filter_by(revision=revision,partnumber=partnumber).first()
    #print(partlist)    
    if len (refpart.children)>0: 
        for i in refpart.children:
            partlist.append((i.getchild(),i.qty*qty))
            get_tree(i.child.partnumber,i.child.revision,partlist,qty=i.qty*qty)
        
    return partlist



def get_flatbom(partnumber,revision,qty=1):
    refpart=Part.query.filter_by(revision=revision,partnumber=partnumber).first()
    flatlist=[]
    flatlist.append((refpart,qty))
    
    dictlist=[]
    checklist=[]
    
    
    get_tree(partnumber,revision,flatlist)
    
    for part,part_qty in flatlist:
        part_dict={}
        part_dict=part.__dict__.copy()
        # part_dict.pop('_sa_instance_state')
        
        if part_dict['partnumber']+part_dict['revision'] in checklist:
            ref_index=checklist.index(part_dict['partnumber']+part_dict['revision'])
            dictlist[ref_index]['totalqty']+=part_qty 
        else:
            part_dict['totalqty']=part_qty  
            dictlist.append(part_dict)
            checklist.append(part_dict['partnumber']+part_dict['revision'])
        
    #print(dictlist)
    
    flatbom=pd.DataFrame(dictlist)
    flatbom['pdfpath']=deliverables_folder+"pdf\\"+flatbom['file']+"_REV_"+flatbom['revision']+".pdf"
    flatbom['pngpath']=deliverables_folder+"png\\"+flatbom['file']+"_REV_"+flatbom['revision']+".png"
    
        
    
    return flatbom
    


################
##### POLISH THE DATABASE CREATION
#db.create_all()





#Job class definition

class Job(db.Model):
    # Defines the Table Name user
    __tablename__ = "job"
    
	# Makes three columns into the table id, name, email
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    jobnumber = db.Column(db.String, nullable=False,unique=True)
    description = db.Column(db.String)
    customer =db.Column(db.String)
    user_id =db.Column(db.Integer, ForeignKey('users.id') , nullable=False)
    date_create=db.Column(db.DateTime, index=True, default=datetime.utcnow)
    date_due=db.Column(db.DateTime)
    date_modify=db.Column(db.DateTime)
    date_finish=db.Column(db.DateTime)
    #scope=relationship('Jobbom',                         
                        # backref='job', 
                        # lazy='dynamic')
    #user=relationship('User', backref = "job")


    def __init__(self, id="",jobnumber="",description="",user="",customer="",user_id="",date_create="",date_due="",date_modify="",date_finish="", **kwargs):
        self.jobnumber=jobnumber
        self.description=description
        self.customer=customer
        self.user_id=user_id
        self.user=user
        self.misc="wahtever"
        print('walksjad;f')
        # self.date_create=datetime.now()
        # self.date_due=date_due
        # self.date_modify=date_modify
        # self.date_finish=date_finish




class Jobbom(db.Model):
    # Defines the Table Name user
    __tablename__ = "jobbom"
    
	# Makes three columns into the table id, name, email
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    job_id = db.Column(db.Integer, ForeignKey('job.id') , nullable=False)
    part_id = db.Column(db.Integer, ForeignKey('part.id') , nullable=False)
    #user_id = db.Column(db.Integer, ForeignKey('users.id') , nullable=False)
    qty =db.Column(db.Integer)

    def __init__(self, job_id="",part_id="",user_id="",qty="", **kwargs):
        self.job_id=job_id
        self.part_id=part_id
        self.user_id=user_id
        self.qty=qty

